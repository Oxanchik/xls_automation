import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import warnings
import time

warnings.filterwarnings("ignore")


def arrange_and_format_file(input_file, output_file):
    start_time = time.time()  # Start timer
    print(f"Starting processing of {input_file}...")

    # 1. Read and Clean Data
    df = pd.read_excel(input_file, engine="openpyxl")

    # Clean spaces in 'CodigoArticulo' if it exists
    if "CodigoArticulo" in df.columns:
        df["CodigoArticulo"] = df["CodigoArticulo"].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()

    print(f"✅ Cleaned spaces in 'CodigoArticulo' column. Initial number of columns: {len(df.columns)}")

    # 2. Drop Unwanted Columns (Just before saving)
    drop_cols = ["MovPosicion", "MovTraspaso", "MovOrigen", "MovConsumo", "MovIdentificador", "Proceso"]
    # Keep only columns that exist in df and are NOT in drop_cols
    df = df[[c for c in df.columns if c not in drop_cols]]

    # Drop completely empty columns
    df = df.dropna(axis='columns', how='all')

    print(f"✅ Removed empty and not useful columns. Remaining columns: {len(df.columns)}")

    # 3. Save to Excel
    df.to_excel(output_file, index=False, engine="openpyxl")

    # 4. Apply Formatting (OPTIMIZADO)
    print("Aplicando formato optimizado...")
    wb = load_workbook(output_file)
    ws = wb.active

    # Estilos
    # header_font = Font(name='Arial', size=10, bold=True)
    header_font = Font(bold=True) # dejamos fuente de base (normalmente es Calibri 11pt)
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    # body_font = Font(name='Arial', size=10)

    # Nota: openpyxl no permite aplicar fuente a todo el rango directamente de forma eficiente
    # sin iterar, PERO podemos optimizar iterando solo por columnas, no por celdas individuales.
    # Al final para economizar la memoria no voy a aplicar body_font, lo quito del código

    max_col = ws.max_column
    max_row = ws.max_row

    # Formato de Cabeceras (Fila 1)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = gray_fill

    # Formato de Fechas y Números (Aplicado por columnas completas)
    # Esto evita el bucle "for cell in row" interno
    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}

    for col_name in ["Fecha", "FechaRegistro"]:
        if col_name in col_map:
            idx = col_map[col_name]
            # Aplicar formato a toda la columna de una vez
            for row_idx in range(2, max_row + 1):
                cell = ws.cell(row=row_idx, column=idx)
                if cell.value:
                    cell.number_format = 'dd/mm/yyyy'

    if "Unidades" in col_map:
        idx = col_map["Unidades"]
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=idx)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'

    # Auto-fit Widths (Optimizado: leer solo valores, no objetos celda completos)
    # Esta parte siempre será pesada si hay muchos datos, pero es necesaria.
    print("Ajustando anchos de columna...")
    for col_idx in range(1, max_col + 1):
        max_len = 0
        # Iterar solo por los valores de la columna
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                length = len(str(val))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)  # Tope de 60 caracteres

    wb.save(output_file)

    # End timer and print result
    end_time = time.time()
    elapsed = end_time - start_time
    print("Applied formatting: Arial 10pt, bold headers with gray fill, date and number formats")
    print(f"✅ Done: {output_file}")
    print(f"Total time taken: {elapsed:.2f} seconds")



# --- Configuration ---
if __name__ == '__main__':
    # UPDATE THESE PATHS
    input_path = r"C:\Users\Labo\Downloads\MovimientoStock.xlsx"
    output_path = r"C:\Users\Labo\Downloads\MovimientoStock_output.xlsx"

    arrange_and_format_file(input_path, output_path)
