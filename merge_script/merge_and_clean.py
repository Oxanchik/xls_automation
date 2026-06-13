import pandas as pd
from pathlib import Path
import warnings
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Suppress the openpyxl default style warning
warnings.filterwarnings("ignore", message="Workbook contains no default style")

def merge_xlsx_files(folder_path="Temp", output_file="Merged.xlsx"):
    """
    Merge all xlsx files in a folder into one file, excluding headers from all except the first file.
    - Adds 'Lote trazado' column with source filename
    - Skips only the header row (first row) from all files except the first
    - Applies Arial 10pt font, bold headers with gray fill
    - Formats date column (A) as dd/mm/yyyy
    - Formats number columns (G, J) with 2 decimals

    Args:
        folder_path: Path to the folder containing xlsx files
        output_file: Name of the output merged file
    """
    # Get all xlsx files in the folder, sorted alphabetically
    xlsx_files = sorted(Path(folder_path).glob("*.xlsx"))

    if not xlsx_files:
        print(f"No xlsx files found in {folder_path}")
        return

    print(f"Found {len(xlsx_files)} xlsx files to merge:")
    for f in xlsx_files:
        print(f"  - {f.name}")

    frames = []

    for i, file_path in enumerate(xlsx_files):
        # Read the Excel file - skip=0 means read all rows (including headers)
        df = pd.read_excel(file_path, engine="openpyxl")

        # Add the 'Lote trazado' column with the filename (without .xlsx extension)
        df["Lote trazado"] = file_path.stem  # removes .xlsx extension
        print(f"Added 'Lote trazado' column with value: {file_path.stem}")

        # Include ALL rows for ALL files
        # Headers are already parsed as column names by pandas, not in the DataFrame
        frames.append(df)
        print(f"Added file: {file_path.name} with {len(df)} data rows")

    # Concatenate all DataFrames
    merged_df: pd.DataFrame = pd.concat(frames, ignore_index=True, sort=False)

    # Remove columns where all values are empty (NaN)
    merged_df = merged_df.dropna(axis='columns', how='all')
    print(f"Removed empty columns. Remaining columns: {len(merged_df.columns)}")

    # Clean excessive spaces in column E (product codes) - 5th column
    merged_df.iloc[:, 4] = merged_df.iloc[:, 4].astype(str).str.strip()

    # Save to Excel file
    merged_df.to_excel(output_file, index=False, engine="openpyxl")
    print(f"\nSuccessfully merged all files into {output_file}")
    print(f"Total rows in merged file: {len(merged_df)}")

    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    arial_10 = Font(name='Arial', size=10)
    bold_arial_10 = Font(name='Arial', size=10, bold=True)
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # Apply Arial 10pt to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = arial_10

    # Apply bold and gray fill to headers in ALL columns that have data (detected automatically)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_arial_10
        cell.fill = gray_fill

    # Apply date format to column A (first column) - dd/mm/yyyy
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value is not None:
            cell.number_format = 'dd/mm/yyyy'

    # Apply number format with 2 decimals to columns G (7th) and J (10th)
    for col_idx in [7, 10]:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = '0.00'

    # Apply text format to column E in the Excel file
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=5)  # Column E = 5
        if cell.value is not None:
            cell.number_format = '@'  # '@' is Excel's text format

    # Auto-fit column widths to content
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                # CRITICAL FIX: Calculate length based on how Excel will display it
                if col_idx == 1 and cell.number_format == 'dd/mm/yyyy':
                    # Force the length to match the format "dd/mm/yyyy" (10 chars)
                    # Or format the value explicitly if it's a datetime object
                    if isinstance(cell.value, datetime):
                        value_str = cell.value.strftime('%d/%m/%Y')
                    else:
                        # Fallback for serial numbers or strings already in date format
                        value_str = "dd/mm/yyyy"  # Safe default length 10
                else:
                    value_str = str(cell.value)

                max_length = max(max_length, len(value_str))
                
        # Set column width with padding
        col_letter = get_column_letter(col_idx)  # Converts 1 -> 'A', 2 -> 'B', etc.
        ws.column_dimensions[col_letter].width = max_length + 2   # This adds 2 characters of padding to each column width for better readability

    # Save the formatted workbook
    wb.save(output_file)
    print("Applied formatting: Arial 10pt, bold headers with gray fill, date and number formats")


# Run the function
if __name__ == '__main__':
    merge_xlsx_files(r"C:\Users\Labo\Downloads", r"C:\Users\Labo\Downloads\Merged.xlsx")

