import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import os

def analyze_loss_by_batch(input_file, output_file):
    """
    Analiza las mermas por lote de producto terminado.
    Agrupa primero para evitar duplicidad de 'Peso_PT'.
    Lee la hoja 'uso_MP', calcula pesos y genera la hoja 'mermas_lotes_PT'.
    En la hoja 'uso_MP' debe haber campo 'Peso_nominal'.

    Regla: Si hay Peso_nominal, se multiplica por Unidades_PT.
           Si no hay Peso_nominal, se toma el valor de Unidades_PT.

    Args:
        input_file (str): Ruta del archivo Excel de entrada.
        output_file (str): Ruta del archivo Excel de salida (puede ser la misma).

    """

    # 1. Cargar datos
    df = pd.read_excel(input_file, sheet_name='uso_MP')

    # 2. Preparar columna de Peso_MP válido (solo KGS, KGR, K, KGRS)
    df['UnidadMedida_MP_norm'] = df['UnidadMedida_MP'].astype(str).str.upper().str.strip()
    unidades_validas = ['KGS', 'KGR', 'K', 'KGRS']

    print("Procesando Materias Primas...")

    # Creamos una columna temporal: si la unidad es válida, tomamos Unidades_MP, si no, 0
    df['Peso_MP_Valido'] = df.apply(
        lambda row: row['Unidades_MP'] if row['UnidadMedida_MP_norm'] in unidades_validas else 0,
        axis=1
    )
    df['Peso_MP_Valido'] = df['Peso_MP_Valido'].fillna(0)

    # 3. Agrupar ANTES de calcular pesos totales
    # Claves: CodigoArticulo_PT, Partida_PT
    group_cols = ['CodigoArticulo_PT', 'Partida_PT']

    agg_dict = {
        'Peso_nominal': 'first',  # Tomamos el valor único (es igual en todas las filas del grupo)
        'Unidades_PT': 'first',  # Tomamos el valor único (evita sumarlas varias veces)
        'UnidadMedida_PT': 'first',  # Dato descriptivo único
        'DescripcionArticulo_PT': 'first',
        'Comentario': 'first',
        'Peso_MP_Valido': 'sum'  # SUMAMOS todas las materias primas válidas del lote
    }

    df_resumen = df.groupby(group_cols, as_index=False).agg(agg_dict)
    df_resumen.rename(columns={'Peso_MP_Valido': 'Peso_MP'}, inplace=True)

    # 4. Calcular Peso_PT TOTAL AHORA (una vez por grupo, sin duplicados)
    # Si Peso_nominal TIENE DATOS -> Multiplicar por Unidades_PT
    # Si Peso_nominal ESTÁ VACÍO/NULL -> Tomar solo Unidades_PT
    print("Calculando Peso_PT...")
    pesos_pt = []
    for index, row in df_resumen.iterrows():
        if pd.notna(row['Peso_nominal']) and row['Peso_nominal'] != 0:
            val = row['Peso_nominal'] * row['Unidades_PT']
        else:
            val = row['Unidades_PT'] if pd.notna(row['Unidades_PT']) else 0
        pesos_pt.append(val)
    df_resumen['Peso_PT'] = pesos_pt

    # 5. Calcular Merma (%)
    # Fórmula: (Peso_MP - Peso_PT) / Peso_MP
    print("Calculando Merma (%)...")

    # Aseguramos que los datos sean numéricos para evitar errores
    peso_mp = df_resumen['Peso_MP'].astype(float)
    peso_pt = df_resumen['Peso_PT'].astype(float)

    # Fórmula vectorizada: (Peso_MP - Peso_PT) / Peso_MP
    # Usamos .replace() para convertir infinitos (división por 0) en 0
    df_resumen['Merma'] = ((peso_mp - peso_pt) / peso_mp).replace([float('inf'), float('-inf')], 0).fillna(0)

    # 6. Limpieza y renombrado de columnas específicas
    # Renombrar Comentario -> Fabricación
    df_resumen.rename(columns={'Comentario': 'Fabricación'}, inplace=True)

    # Eliminar el prefijo "Fabricación: " (case-insensitive y espacios extra)
    # Usamos regex=False para tratarlo como texto plano, o regex=True si hay variaciones complejas
    if 'Fabricación' in df_resumen.columns:
        df_resumen['Fabricación'] = df_resumen['Fabricación'].astype(str).str.replace(
            'Fabricación:', '', case=False, regex=False
        ).str.strip()

        # Opcional: Si el resultado es "nan" o vacío por valores nulos originales, dejarlo vacío limpio
        df_resumen['Fabricación'] = df_resumen['Fabricación'].replace('nan', '')

    # 7. Reordenar columnas explícitamente
    # Orden solicitado: Codigo, Descripcion, Partida, Fabricación, ...resto..., Peso_PT, Merma
    columnas_finales = [
        'CodigoArticulo_PT',
        'DescripcionArticulo_PT',
        'Partida_PT',
        'Fabricación',
        'UnidadMedida_PT',
        'Peso_nominal',
        'Unidades_PT',
        'Peso_MP',
        'Peso_PT',
        'Merma'
    ]

    # Filtramos por si alguna columna opcional no existiera, evitando errores
    columnas_finales = [c for c in columnas_finales if c in df_resumen.columns]
    df_resumen = df_resumen[columnas_finales]

    # 8. Ordenar la tabla
    df_resumen = df_resumen.sort_values(
        by=['CodigoArticulo_PT', 'Partida_PT'],
        ascending=[True, True]
    )

    # 9. Guardar y aplicar formato con OpenPyXL
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # --- FASE A: ESCRITURA DE DATOS ---
            df_resumen.to_excel(writer, sheet_name='mermas_lotes_PT', index=False)

            # --- FASE B: FORMATEO ---
            worksheet = writer.sheets['mermas_lotes_PT']

            # Definir estilos (Azul Excel #4472C4, Texto Blanco, Negrita)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 1. Formatear Cabecera (Fila 1)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 2. Identificar el índice de la columna 'Merma' (0-based)
            if 'Merma' in df_resumen.columns:
                merma_idx = df_resumen.columns.get_loc('Merma')
                merma_col_letter = get_column_letter(merma_idx + 1)

                # Aplicar formato de porcentaje a todas las celdas de datos (desde fila 2)
                # worksheet[f'{col_letter}2:{col_letter}{last_row}']
                last_row = len(df_resumen) + 1

                for row_idx in range(2, last_row + 1):
                    cell = worksheet[f'{merma_col_letter}{row_idx}']
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right')

            # 3. Ajustar anchos de columna automáticamente
            for i, col in enumerate(df_resumen.columns):
                # Calcular longitud máxima del contenido
                data_series = df_resumen[col].fillna('').astype(str)
                max_data_len = int(data_series.map(len).max()) if len(data_series) > 0 else 0
                header_len = len(col)

                # Margen de seguridad +2 y límite máximo de 50
                max_len = max(max_data_len, header_len) + 2
                col_width = min(max_len, 50)

                # Aplicar ancho
                col_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[col_letter].width = col_width

        print(f"\nÉxito: Hoja 'mermas_lotes_PT' generada, ordenada y formateada en {output_file}")

    except Exception as e:
        print(f"Error al guardar o formatear el Excel: {e}")
        raise

if __name__ == '__main__':
    input_p = r"C:\Users\Labo\Downloads\MovimientoStock_output.xlsx"
    if os.path.exists(input_p):
        analyze_loss_by_batch(input_p, input_p)
    else:
        print("Archivo no encontrado.")
