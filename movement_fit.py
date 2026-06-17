import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import os


def find_column(df, target_name):
    """Busca una columna ignorando mayúsculas y espacios extra."""
    target_clean = target_name.lower().strip()
    for col in df.columns:
        if col.lower().strip() == target_clean:
            return col
    # Búsqueda parcial si no hay exacta
    for col in df.columns:
        if target_clean in col.lower():
            return col
    return None


def analyze_material_usage(input_file, output_file):
    print(f"🔍 Iniciando análisis de uso de materia prima en: {os.path.basename(input_file)}")

    carpeta_base = os.path.dirname(os.path.abspath(input_file))

    # 1. Cargar datos
    df = pd.read_excel(input_file, engine="openpyxl", dtype={'CodigoArticulo': str, 'Partida': str})

    # --- Limpieza de nombres de columnas ---
    df.columns = df.columns.str.strip()

    # Buscar columna Comentario de forma flexible
    col_comment_real = find_column(df, "Comentario")
    if not col_comment_real:
        print("❌ Error: No se encontró la columna 'Comentario'.")
        return

    # Renombrar a "Comentario" para estandarizar
    if col_comment_real != "Comentario":
        df.rename(columns={col_comment_real: "Comentario"}, inplace=True)

    # Nombres de columnas estandarizados
    col_type = "TipoMovimiento"
    col_origin = "OrigenMovimiento"
    col_code = "CodigoArticulo"
    col_batch = "Partida"
    col_desc = "DescripcionArticulo"
    col_qty = "Unidades"
    col_unit = "UnidadMedida1_"

    # Validar columnas esenciales
    required_cols = ["Comentario", col_type, col_origin, col_code, col_qty, col_desc, col_unit]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        print(f"❌ Faltan columnas esenciales: {missing}")
        return

    # 2. Identificar Productos Terminados (PT) y Materia Prima (MP)
    df_f = df[df[col_origin] == 'F'].copy()

    if df_f.empty:
        print("⚠️ No se encontraron movimientos de tipo 'F' (Fabricación).")
        return

    df_pt = df_f[df_f[col_type] == 1].copy()  # Entrada = PT
    df_mp = df_f[df_f[col_type] == 2].copy()  # Salida = MP

    print(f"   - Filas de Fabricación (F): {len(df_f)}")
    print(f"   - Productos Terminados (Entrada/1): {len(df_pt)}")
    print(f"   - Materia Prima (Salida/2): {len(df_mp)}")

    if df_pt.empty or df_mp.empty:
        print("⚠️ No hay suficientes datos para cruzar.")
        return

    # 3. Agrupar Datos

    # pt_summary = df_pt.groupby([col_code, col_batch, "Comentario"]).agg(
    #     DescripcionArticulo=(col_desc, 'first'),
    #     CantidadProducida=(col_qty, 'sum'),
    #     UnidadMedida=(col_unit, 'first')
    # ).reset_index()
    #
    # Antes de reset_index(): Las columnas col_code, col_batch y "Comentario" son el índice del DataFrame (no columnas regulares).
    # Si omites .reset_index(), ocurrirá lo siguiente:
    # 1.	Las columnas agrupadas permanecen como índice: No podrás acceder a col_code, col_batch o "Comentario" como columnas normales usando df["columna"].
    # 2.	Problemas al exportar: Al guardar el DataFrame a Excel o CSV, el índice se guardará de forma diferente, posiblemente creando columnas no deseadas o perdiendo la estructura esperada.
    # 3.	Dificultad para operaciones posteriores: Si intentas hacer merges, filtros o accesos por nombre de columna, fallará porque esas columnas están en el índice, no en las columnas del DataFrame.
    # 4.	Posible error al escribir a Excel: La función dataframe_to_rows de openpyxl podría no manejar correctamente un DataFrame con MultiIndex, generando un formato inesperado en el archivo de salida.

    # Al usar as_index=False directamente dentro de groupby(), le indicas a pandas que no convierta las columnas de agrupación en el índice,
    # sino que las mantenga como columnas normales desde el principio.

    # Resumen PT: Agrupar por Código, Partida y Comentario
    pt_summary = df_pt.groupby([col_code, col_batch, "Comentario"], as_index=False).agg(
        DescripcionArticulo_PT=(col_desc, "first"),
        Unidades_PT=(col_qty, "sum"),
        UnidadMedida_PT=(col_unit, "first")
    )

    # Resumen MP: Agrupar por Código, Partida y Comentario
    # IMPORTANTE: Incluimos también Descripción y Unidad en la agrupación o las tomamos después
    mp_summary = df_mp.groupby([col_code, col_batch, "Comentario"], as_index=False).agg(
        DescripcionArticulo_MP=(col_desc, "first"),
        Unidades_MP=(col_qty, "sum"),
        UnidadMedida_MP=(col_unit, "first")
    )

    # 4. Cruzar Datos (Merge)
    # how = "inner" -> hace que aparezcan en la hoja "uso_MP" solo PT / MP que coinciden.
    # how = "left" -> hace que aparezcan en la hoja "uso_MP" también PT que no tienen "pareja" en MP.
    merged = pd.merge(
        pt_summary,
        mp_summary,
        on="Comentario",
        how="left",
        suffixes=("_PT", "_MP")
    )

    if merged.empty:
        print("⚠️ El cruce de datos no produjo resultados.")
        return

    # 5. Crear Tabla Normal (Plana)
    final_table = merged[[
        "CodigoArticulo_PT",
        "Partida_PT",
        "DescripcionArticulo_PT",
        "Unidades_PT",
        "UnidadMedida_PT",
        "CodigoArticulo_MP",
        "Partida_MP",
        "DescripcionArticulo_MP",
        "Unidades_MP",
        "UnidadMedida_MP",
        "Comentario"
    ]].copy()

    # --- AGREGAR PESO NOMINAL ---

    # A. Cargar el archivo de pesos
    pesos_file = os.path.join(carpeta_base, "Lista_pesos.xlsx")

    if not os.path.exists(pesos_file):
        print(f"❌ Error: No se encuentra el archivo de pesos '{pesos_file}' en la carpeta actual.")
        return

    try:
        df_pesos = pd.read_excel(
            pesos_file,
            sheet_name="PT_UDS",
            dtype={'CodigoArticulo': str, 'Peso_nominal_kg': float}  # Forzar tipos correctos
        )
        # Limpiar nombres de columnas por seguridad
        df_pesos.columns = df_pesos.columns.str.strip()

        # Validar columnas en archivo de pesos
        if 'CodigoArticulo' not in df_pesos.columns or 'Peso_nominal_kg' not in df_pesos.columns:
            print("❌ Error: El archivo de pesos no tiene las columnas 'CodigoArticulo' o 'Peso_nominal_kg'.")
            return

    except Exception as e:
        print(f"❌ Error al leer el archivo de pesos: {e}")
        return

    # B. Fusionar con la tabla final
    # Unimos usando CodigoArticulo_PT del resultado y CodigoArticulo de la lista de pesos
    final_table = pd.merge(
        final_table,
        df_pesos[['CodigoArticulo', 'Peso_nominal_kg']],
        left_on='CodigoArticulo_PT',
        right_on='CodigoArticulo',
        how='left'
    )

    # Renombrar para que quede limpio y eliminar la columna duplicada de código
    final_table.rename(columns={'Peso_nominal_kg': 'Peso_nominal'}, inplace=True)
    final_table.drop(columns=['CodigoArticulo'], inplace=True)

    # C. Reordenar columnas para poner 'Peso_nominal' justo después de 'UnidadMedida_PT'
    columnas_ordenadas = [
        "CodigoArticulo_PT",
        "Partida_PT",
        "DescripcionArticulo_PT",
        "Unidades_PT",
        "UnidadMedida_PT",
        "Peso_nominal",  # Nueva columna insertada aquí
        "CodigoArticulo_MP",
        "Partida_MP",
        "DescripcionArticulo_MP",
        "Unidades_MP",
        "UnidadMedida_MP",
        "Comentario"
    ]

    # Asegurar que todas las columnas existan antes de reordenar
    columnas_finales = [c for c in columnas_ordenadas if c in final_table.columns]
    final_table = final_table[columnas_finales]

    # 6. Guardar en nueva hoja "uso_MP"
    try:
        # 1. Abrir el escritor (esto carga el libro si mode='a')
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

            # --- FASE A: ESCRITURA RÁPIDA DE DATOS ---
            final_table.to_excel(writer, sheet_name="uso_MP", index=False)

            # --- FASE B: FORMATEO CON OPENPYXL ---
            # Accedemos directamente al objeto workbook y worksheet que pandas acaba de crear/modificar
            # workbook = writer.book
            worksheet = writer.sheets["uso_MP"]

            # Definir estilos (Azul Excel #4472C4, Texto Blanco, Negrita)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center")

            # 1. Formatear Cabecera (Fila 1)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 2. Ajustar anchos de columna automáticamente (basado en el contenido)
            # Iteramos sobre las columnas del DataFrame para calcular el ancho ideal
            for i, col in enumerate(final_table.columns):
                # Calcular longitud máxima: entre la cabecera y el dato más largo
                # Rellenar NaN con cadena vacía antes de convertir a string y medir
                data_series = final_table[col].fillna('').astype(str).map(len)
                max_val = data_series.max()

                # Asegurar que es un entero (por si la serie estuviera vacía)
                max_data_len = int(max_val) if pd.notna(max_val) else 0
                header_len = len(col)

                max_len = max(max_data_len, header_len) + 2 # + 2 es un margen de seguridad

                # Limitar ancho máximo para que no sea ridículo (ej. 50 caracteres)
                col_width = min(max_len, 50)

                col_letter = get_column_letter(i + 1)  # Convierte 1 -> 'A', 2 -> 'B', etc.
                worksheet.column_dimensions[col_letter].width = col_width

            # 3. Ejemplo: Formato condicional o específico para la columna 'Peso_nominal'
            # Supongamos que 'Peso_nominal' es la columna F (índice 5)
            # Podrías aplicar formato de número aquí si fuera necesario

        print(f"✅ Hoja 'uso_MP' creada y formateada exitosamente con {len(final_table)} registros.")

    except PermissionError:
        print("❌ Error: No se pudo guardar. Asegúrate de que el archivo Excel esté CERRADO.")
    except Exception as e:
        print(f"❌ Error crítico: {e}")


if __name__ == '__main__':
    input_p = r"C:\Users\Labo\Downloads\MovimientoStock_output.xlsx"
    if os.path.exists(input_p):
        analyze_material_usage(input_p, input_p)
    else:
        print("Archivo no encontrado.")
